from uuid import uuid4
from datetime import timedelta, datetime, date
from calendar import monthrange
import openpyxl
import json
import requests
import os
from copy import copy

from src.route.services.route_core import StepType, DurationItem, TimesheetDayType
from src.message_broker.producer import send_into_timesheet_queue

MONTH_NAMES = {1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
               7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'}


def create_timesheet_for_new_trainee(user_id, route):
    steps_count = 0
    steps_by_order_dict = {}
    for step in route.steps:
        if step.step_type == StepType.timeSheet:
            steps_count += 1
            steps_by_order_dict[step.order] = {
                "duration": step.time_sheep_step[0].duration,
                "duration_item": step.time_sheep_step[0].duration_item,
                "minimum_fill_percent": step.time_sheep_step[0].minimum_fill_percent,
            }
    start_date = route.start_date
    for i in range(1, steps_count + 1):
        step_info = steps_by_order_dict[i]
        if step_info["duration_item"] == DurationItem.months:
            step_info["duration"] = step_info["duration"] * 30
        generate_timesheet_data = {
            "eventId": str(uuid4()),
            "userId": str(user_id),
            "routeId": str(route.id),
            "eventType": "generate",
            "toGenerate": {
                "dateStart": str(start_date),
                "duration": step_info["duration"],
                "durationItem": DurationItem.days
            }
        }
        start_date = start_date + timedelta(days=step_info["duration"])
        send_into_timesheet_queue(json.dumps(generate_timesheet_data))


def check_timesheet_day_belongs_to_user(user_id, route_id, timesheet_day_id, timesheet_date):
    url = f'{os.getenv("TIMESHEET_URL")}/api/v1/timesheet/{user_id}/'
    user_timesheet = requests.get(url, params={"routeId": route_id})
    for timesheet_day in user_timesheet.json():
        if timesheet_day_id == timesheet_day["id"] and timesheet_date == timesheet_day["timeSheetDate"]:
            return True
    return False


def prepare_info_to_update_timesheet_day(user_id, route_id, timesheet_date, day_type):
    update_timesheet_data = {
        "eventId": str(uuid4()),
        "userId": str(user_id),
        "routeId": str(route_id),
        "eventType": "update",
        "toUpdate": {
            "eventDate": timesheet_date,
            "dayType": day_type
        },
    }
    return update_timesheet_data


def get_long_name_for_date_interval(start_date, end_date):
    return f"с {start_date.day} по {end_date.day} {MONTH_NAMES[start_date.month]} {start_date.year} г."


def get_short_name_for_date_interval(start_date, end_date):
    return f"{start_date.day}-{end_date.day} {MONTH_NAMES[start_date.month]} {start_date.year}"


def get_timesheet_intervals(start_date, end_date, today):
    days_in_start_month = monthrange(start_date.year, start_date.month)[1]
    last_day_of_start_month = date(start_date.year, start_date.month, days_in_start_month)
    if today <= last_day_of_start_month:
        if today < end_date:
            return {1: {
                "start_date": start_date, "end_date": today,
                "long_name": get_long_name_for_date_interval(start_date, today),
                "short_name": get_short_name_for_date_interval(start_date, today)
            }}
        else:
            return {1: {
                "start_date": start_date, "end_date": end_date,
                "long_name": get_long_name_for_date_interval(start_date, end_date),
                "short_name": get_short_name_for_date_interval(start_date, end_date)
            }}
    else:
        if end_date < last_day_of_start_month:
            return {1: {
                "start_date": start_date, "end_date": end_date,
                "long_name": get_long_name_for_date_interval(start_date, end_date),
                "short_name": get_short_name_for_date_interval(start_date, end_date)
            }}
    timesheet_intervals = {1: {
        "start_date": start_date, "end_date": last_day_of_start_month,
        "long_name": get_long_name_for_date_interval(start_date, last_day_of_start_month),
        "short_name": get_short_name_for_date_interval(start_date, last_day_of_start_month)
    }}

    interval_number = 1
    if start_date.month + 1 <= 12:
        first_day_next_month = date(start_date.year, start_date.month + 1, 1)
    else:
        first_day_next_month = date(start_date.year + 1, 1, 1)
    while first_day_next_month <= end_date and first_day_next_month <= today:
        interval_number += 1
        days_in_month = monthrange(first_day_next_month.year, first_day_next_month.month)[1]
        last_day_of_month = date(first_day_next_month.year, first_day_next_month.month, days_in_month)
        if today <= last_day_of_month:
            if today < end_date:
                timesheet_intervals[interval_number] = {
                    "start_date": first_day_next_month, "end_date": today,
                    "long_name": get_long_name_for_date_interval(first_day_next_month, today),
                    "short_name": get_short_name_for_date_interval(first_day_next_month, today)
                }
            else:
                timesheet_intervals[interval_number] = {
                    "start_date": first_day_next_month, "end_date": end_date,
                    "long_name": get_long_name_for_date_interval(first_day_next_month, end_date),
                    "short_name": get_short_name_for_date_interval(first_day_next_month, end_date)
                }
        else:
            if end_date < last_day_of_month:
                timesheet_intervals[interval_number] = {
                    "start_date": first_day_next_month, "end_date": end_date,
                    "long_name": get_long_name_for_date_interval(first_day_next_month, end_date),
                    "short_name": get_short_name_for_date_interval(first_day_next_month, end_date)
                }
            else:
                timesheet_intervals[interval_number] = {
                    "start_date": first_day_next_month, "end_date": last_day_of_month,
                    "long_name": get_long_name_for_date_interval(first_day_next_month, last_day_of_month),
                    "short_name": get_short_name_for_date_interval(first_day_next_month, last_day_of_month)
                }
        if first_day_next_month.month + 1 <= 12:
            first_day_next_month = date(first_day_next_month.year, first_day_next_month.month + 1, 1)
        else:
            first_day_next_month = date(first_day_next_month.year + 1, 1, 1)
    return timesheet_intervals


def get_base_timesheet_for_month(year, month):
    first_weekday_of_month, max_day_of_month = monthrange(year, month)
    first_weekday_of_month = first_weekday_of_month + 1  # ISO, monday = 1, no 0
    base_timesheet_for_month = dict()
    for i in range(1, max_day_of_month + 1):
        if (first_weekday_of_month + i - 1) % 7 in range(1, 6):
            base_timesheet_for_month[i] = TimesheetDayType.work
        else:
            base_timesheet_for_month[i] = TimesheetDayType.dayOff
    return base_timesheet_for_month


def get_trainee_timesheet_by_year_and_month(trainee_timesheet):
    trainee_timesheet_by_year_and_month = dict()
    for timesheet_day in trainee_timesheet:
        day = datetime.strptime(timesheet_day["timeSheetDate"], '%Y-%m-%d')
        if day.year not in trainee_timesheet_by_year_and_month.keys():
            trainee_timesheet_by_year_and_month[day.year] = {}
        if day.month not in trainee_timesheet_by_year_and_month[day.year].keys():
            trainee_timesheet_by_year_and_month[day.year][day.month] = {}
        trainee_timesheet_by_year_and_month[day.year][day.month][day.day] = timesheet_day["dayType"]
    return trainee_timesheet_by_year_and_month


def get_pair_of_day_value_for_timesheet(trainee_day, base_day):
    if base_day == TimesheetDayType.work:
        if trainee_day == TimesheetDayType.work:
            return 8, 'Я'
        elif trainee_day == TimesheetDayType.ill:
            return None, 'Б'
        else:
            return None, 'А'
    else:  # TimesheetDayType.dayOff
        if trainee_day == TimesheetDayType.work:
            return 8, 'Я'
        elif trainee_day == TimesheetDayType.ill:
            return None, 'Б'
        elif trainee_day == TimesheetDayType.dayOff:
            return None, 'В'
        else:  # TimesheetDayType.vacation
            return None, 'А'


def get_trainees_timesheet_for_name_year_and_month(trainees):
    url_to_get_trainee_timesheet = f'{os.getenv("TIMESHEET_URL")}/api/v1/timesheet/'
    trainees_timesheet = dict()
    for trainee in trainees:
        url = url_to_get_trainee_timesheet + str(trainee.user_id) + '/'
        trainee_timesheet = requests.get(url, params={"routeId": trainee.route_id}).json()
        trainee_timesheet_by_year_and_month = get_trainee_timesheet_by_year_and_month(trainee_timesheet)
        trainees_timesheet[trainee.user.name] = trainee_timesheet_by_year_and_month
    return trainees_timesheet


def fill_sheet(sheet_index, sheet, interval_info, trainees_timesheet):
    sheet.sheet_view.showGridLines = False
    sheet.title = interval_info["short_name"]
    start_date = interval_info["start_date"]
    month = start_date.month
    year = start_date.year
    last_day = interval_info["end_date"].day

    base_timesheet_for_month = get_base_timesheet_for_month(year, month)

    column_offset_for_first_half_of_the_month = 9
    column_offset_for_second_half_of_the_month = 12
    offset_for_trainee_rows = 0
    sheet.cell(row=3, column=22, value=sheet_index+1)
    sheet.cell(row=6, column=17, value=f'за период {interval_info["long_name"]}')
    for trainee_name in trainees_timesheet.keys():
        first_row_for_trainee = 15 + offset_for_trainee_rows
        sheet.cell(row=first_row_for_trainee, column=1, value=offset_for_trainee_rows / 2 + 1)
        sheet.cell(row=first_row_for_trainee, column=3, value=trainee_name)
        sheet.cell(row=first_row_for_trainee, column=4, value='Куратор')
        sheet.cell(row=first_row_for_trainee + 1, column=4, value='Куратор')
        sheet.cell(row=first_row_for_trainee, column=5, value=40)
        sheet.cell(row=first_row_for_trainee, column=8, value='Специалист проекта в сфере городского управления')
        sheet.merge_cells(start_row=first_row_for_trainee, start_column=8, end_row=first_row_for_trainee + 1,
                          end_column=9)
        is_trainee_info_empty = False
        trainee_timesheet = trainees_timesheet[trainee_name]
        if year not in trainee_timesheet.keys():
            is_trainee_info_empty = True
        else:
            if month not in trainee_timesheet[year].keys():
                is_trainee_info_empty = True
            else:
                trainee_timesheet_for_month = trainee_timesheet[year][month]
        for day_of_month in range(1, last_day+1):
            if day_of_month < 16:
                column_offset = column_offset_for_first_half_of_the_month
            else:
                column_offset = column_offset_for_second_half_of_the_month
            if not is_trainee_info_empty:
                if day_of_month in trainee_timesheet_for_month.keys():
                    value_1, value_2 = get_pair_of_day_value_for_timesheet(
                        trainee_timesheet_for_month[day_of_month], base_timesheet_for_month[day_of_month]
                    )
                else:
                    value_1, value_2 = '-', '-'
            else:
                value_1, value_2 = '-', '-'
            sheet.cell(row=first_row_for_trainee, column=day_of_month + column_offset, value=value_1)
            sheet.cell(row=first_row_for_trainee + 1, column=day_of_month + column_offset, value=value_2)
            sheet[f'Y{first_row_for_trainee}'] = f'=COUNT(J{first_row_for_trainee}:X{first_row_for_trainee})'
            sheet[f'Y{first_row_for_trainee + 1}'] = f'=SUM(J{first_row_for_trainee}:X{first_row_for_trainee})'
            sheet[f'AR{first_row_for_trainee}'] = \
                f'=COUNT(J{first_row_for_trainee}:X{first_row_for_trainee})+COUNT(AB{first_row_for_trainee}:AQ{first_row_for_trainee})'
            sheet[f'AR{first_row_for_trainee + 1}'] = \
                f'=SUM(J{first_row_for_trainee}:X{first_row_for_trainee})+SUM(AB{first_row_for_trainee}:AQ{first_row_for_trainee})'
        offset_for_trainee_rows += 2


def export_timesheet_as_excel(filepath, route):
    trainees = route.trainees
    today = datetime.now().date()
    timesheet_intervals = get_timesheet_intervals(route.start_date, route.end_date, today)
    data_to_fill_file = {
        'Учреждение': 'Государственное автономное образовательное учреждение высшего образования "Московский городской университет управления Правительства Москвы имени Ю.М. Лужкова"',
        'Вид табеля': 'первичный',
        'Форма по ОКУД': '0504421',
        'Дата': today.strftime("%d.%m.%Y"),
        'по ОКПО': '40002552',
        'Дата формирования документа': today.strftime("%d.%m.%Y"),
        'Руководитель подразделения': 'Филимонова Н.Ю.',
        'Исполнитель': 'Волкова К.Г.',

    }
    excel_file = openpyxl.load_workbook(filepath)
    sheet = excel_file.active

    sheet.cell(row=7, column=8, value=data_to_fill_file['Учреждение'])
    sheet.cell(row=9, column=8, value=data_to_fill_file['Вид табеля'])
    sheet.cell(row=5, column=46, value=data_to_fill_file['Форма по ОКУД'])
    sheet.cell(row=6, column=46, value=data_to_fill_file['Дата'])
    sheet.cell(row=7, column=46, value=data_to_fill_file['по ОКПО'])
    sheet.cell(row=10, column=46, value=data_to_fill_file['Дата формирования документа'])

    if len(trainees) == 0:
        row_offset = -2
        sheet.delete_rows(15, -row_offset)
    elif len(trainees) == 1:
        row_offset = 0
    else:
        row_offset = (len(trainees) - 1) * 2
        sheet.insert_rows(17, row_offset)
        for col in range(1, 53):
            for row in range(17, 17 + row_offset):
                sheet.cell(row=row, column=col)._style = copy(sheet.cell(row=15, column=col)._style)
        columns_to_merge_for_one_row = [(25, 27), (44, 48)]
        for row in range(16, 16 + row_offset + 1):
            sheet.row_dimensions[row].height = sheet.row_dimensions[15].height
            for start_col, end_col in columns_to_merge_for_one_row:
                sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        multi_row_merge_cells = [(15, 1, 16, 1), (15, 2, 16, 2), (15, 3, 16, 3), (15, 5, 16, 5), (15, 6, 16, 6),
                                 (15, 7, 16, 7), (15, 8, 16, 9)]
        for i in range(2, row_offset + 1, 2):
            for start_row, start_col, end_row, end_col in multi_row_merge_cells:
                sheet.merge_cells(start_row=start_row + i, start_column=start_col, end_row=end_row + i,
                                  end_column=end_col)

    template_file_bottom_part = openpyxl.load_workbook('/application/media/timesheet_template_bottom_part.xlsx')
    template_sheet = template_file_bottom_part.active
    start_bottom_row = 17 + row_offset
    for row in range(1, 11):
        sheet.row_dimensions[row + start_bottom_row - 1].height = template_sheet.row_dimensions[row].height
        for col in range(1, 53):
            sheet.cell(row=row + start_bottom_row - 1, column=col).value = template_sheet.cell(row=row,
                                                                                               column=col).value
            if template_sheet.cell(row=row, column=col).has_style:
                sheet.cell(row=row + start_bottom_row - 1, column=col).font = \
                    copy(template_sheet.cell(row=row, column=col).font)
                sheet.cell(row=row + start_bottom_row - 1, column=col).border = \
                    copy(template_sheet.cell(row=row, column=col).border)
                sheet.cell(row=row + start_bottom_row - 1, column=col).fill = \
                    copy(template_sheet.cell(row=row, column=col).fill)
                sheet.cell(row=row + start_bottom_row - 1, column=col).number_format = \
                    copy(template_sheet.cell(row=row, column=col).number_format)
                sheet.cell(row=row + start_bottom_row - 1, column=col).protection = \
                    copy(template_sheet.cell(row=row, column=col).protection)
                sheet.cell(row=row + start_bottom_row - 1, column=col).alignment = \
                    copy(template_sheet.cell(row=row, column=col).alignment)
    cells_to_merge_in_bottom_part = [[3, 10, 3, 12], [2, 14, 2, 20], [3, 14, 3, 20], [5, 15, 5, 20],
                                     [6, 10, 6, 12], [6, 14, 6, 20], [6, 28, 6, 32], [6, 34, 6, 37],
                                     [6, 39, 6, 45]]
    for s_r, s_c, e_r, e_c in cells_to_merge_in_bottom_part:
        sheet.merge_cells(start_row=s_r + start_bottom_row, start_column=s_c,
                          end_row=e_r + start_bottom_row, end_column=e_c)
    sheet.cell(row=2 + start_bottom_row, column=14, value=data_to_fill_file['Руководитель подразделения'])
    sheet.cell(row=5 + start_bottom_row, column=15, value=data_to_fill_file['Исполнитель'])
    for _ in range(len(timesheet_intervals) - 1):
        excel_file.copy_worksheet(sheet)

    trainees_timesheet = get_trainees_timesheet_for_name_year_and_month(trainees)

    for sheet_index in range(len(timesheet_intervals)):
        excel_file.active = sheet_index
        sheet_to_fill = excel_file.active
        fill_sheet(sheet_index, sheet_to_fill, timesheet_intervals[sheet_index+1], trainees_timesheet)
    excel_file.save(filepath)
